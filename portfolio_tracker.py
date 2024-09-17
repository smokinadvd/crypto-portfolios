import requests
import pandas as pd
import logging
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os

# Setup logging
logging.basicConfig(level=logging.DEBUG)

# Define the API endpoints and the provided API keys
api_key = "8e58e4ac-182a-4c41-ac41-6f7032cfd47c"
url_latest = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest"
url_quotes = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/quotes/latest"
headers = {
    'Accepts': 'application/json',
    'X-CMC_PRO_API_KEY': api_key,
}

alpha_vantage_api_key = "BVGR4LWJ5G0F11HD"
alpha_vantage_url = "https://www.alphavantage.co/query"

portfolio_file = "portfolios.xlsx"

class Coin:
    def __init__(self, id, symbol, name, price, change_24h, change_7d, market_cap, date_added):
        self.id = id
        self.symbol = symbol
        self.name = name
        self.price = price if price is not None else 0.0
        self.change_24h = change_24h if change_24h is not None else 0.0
        self.change_7d = change_7d if change_7d is not None else 0.0
        self.market_cap = market_cap if market_cap is not None else 0.0
        self.date_added = date_added

    def __str__(self):
        return f"{self.symbol} | {self.name} | ${self.price:.8f} | {self.change_24h:.2f}% | {self.change_7d:.2f}% | ${self.market_cap:.2f}"

def get_latest_meme_coins(min_market_cap=100000):
    params = {
        'start': '1',
        'limit': '5000',
        'sort': 'date_added',
        'sort_dir': 'desc',
        'convert': 'USD'
    }
    response = requests.get(url_latest, headers=headers, params=params)
    filtered_meme_coins = []

    if response.status_code == 200:
        data = response.json()
        for coin in data['data']:
            tags = coin.get('tags', [])
            if 'memes' in tags:
                id = coin['id']
                symbol = coin['symbol']
                name = coin['name']
                quote = coin['quote']['USD']
                price = quote.get('price', 0)
                change_24h = quote.get('percent_change_24h', 0)
                change_7d = quote.get('percent_change_7d', 0)
                market_cap = quote.get('market_cap', 0)
                if market_cap >= min_market_cap:
                    date_added = coin.get('date_added', '1970-01-01T00:00:00Z')
                    coin_instance = Coin(id, symbol, name, price, change_24h, change_7d, market_cap, date_added)
                    filtered_meme_coins.append(coin_instance)

                    if len(filtered_meme_coins) == 100:
                        break
    else:
        logging.error(f"Error: {response.json()['status']['error_message']}")
    
    logging.info(f"Number of coins fetched from CoinMarketCap: {len(filtered_meme_coins)}")
    return filtered_meme_coins

def get_alpha_vantage_price(symbol, market):
    url = f'https://www.alphavantage.co/query?function=CURRENCY_EXCHANGE_RATE&from_currency={symbol}&to_currency={market}&apikey={alpha_vantage_api_key}'
    response = requests.get(url)

    if response.status_code == 200:
        data = response.json()
        try:
            price = data['Realtime Currency Exchange Rate']['5. Exchange Rate']
            logging.info(f"The current price for {symbol} in {market} is: ${price}")
            return float(price)
        except KeyError:
            logging.error(f"Error retrieving data for {symbol}-{market}: {data}")
            return None
    else:
        logging.error(f"Error fetching Alpha Vantage data for {symbol}-{market}: {response.status_code}")
        return None

def load_portfolio_data():
    if not os.path.exists(portfolio_file):
        workbook = Workbook()
        workbook.save(portfolio_file)

def save_portfolio_data(portfolio_name, coin_data, index_prices):
    workbook = load_workbook(portfolio_file)
    sheet = workbook.create_sheet(title=portfolio_name)

    # Create header row
    headers = ['ID', 'Symbol', 'Name', 'Price', 'Change 24h', 'Change 7d', 'Market Cap', 'Date Added']
    sheet.append(headers)

    # Append each coin's data to the sheet
    for coin in coin_data:
        row = [
            coin['id'], coin['symbol'], coin['name'], coin['price'], 
            coin['change_24h'], coin['change_7d'], coin['market_cap'], coin['date_added']
        ]
        sheet.append(row)

    # Add BTC and SOL prices in the next rows for reference
    sheet.append([])
    sheet.append(['BTC Price', index_prices['BTC']])
    sheet.append(['SOL Price', index_prices['SOL']])

    workbook.save(portfolio_file)

def create_new_portfolio():
    load_portfolio_data()
    workbook = load_workbook(portfolio_file)

    # Determine the new portfolio name
    portfolio_count = len(workbook.sheetnames)
    portfolio_name = f"Portfolio_{portfolio_count + 1}"

    latest_meme_coins = get_latest_meme_coins()
    coin_data = [{
        'id': coin.id,
        'symbol': coin.symbol,
        'name': coin.name,
        'price': coin.price,
        'change_24h': coin.change_24h,
        'change_7d': coin.change_7d,
        'market_cap': coin.market_cap,
        'date_added': coin.date_added
    } for coin in latest_meme_coins]

    index_prices = {
        'BTC': get_alpha_vantage_price('BTC', 'USD'),
        'SOL': get_alpha_vantage_price('SOL', 'USD')
    }

    save_portfolio_data(portfolio_name, coin_data, index_prices)

def update_portfolios():
    workbook = load_workbook(portfolio_file)
    today = datetime.now()

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Check if the portfolio needs to be updated (only if older than 30 days)
        creation_date = sheet.cell(row=1, column=1).value
        if today >= creation_date + timedelta(days=30):
            # Update the portfolio with new BTC and SOL prices
            btc_price = get_alpha_vantage_price('BTC', 'USD')
            sol_price = get_alpha_vantage_price('SOL', 'USD')

            # Append new price data at the end of the sheet
            sheet.append([])
            sheet.append(['BTC Price Update', btc_price])
            sheet.append(['SOL Price Update', sol_price])

    workbook.save(portfolio_file)

if __name__ == "__main__":
    create_new_portfolio()
    update_portfolios()
